import openpyxl
from datetime import datetime
import os

# book = openpyxl.load_workbook("C:\\Users\\satheeshnair\\PycharmProjects\\3PI\\testdata\\testdata.xlsx")
# sheet = book.active
# rows = sheet.max_row
# cols = sheet.max_column
#
# for col in range(1,cols+1):
#     for row in range(2,rows+1):
#         testcasename = sheet.cell(row=row,column=col).value
#         if (testcasename == "test_shipto"):
#             methodname = sheet.cell(row=row,column=col+1).value
#             if(methodname == "Expected_date"):
#                 value = sheet.cell(row=row,column=col+2).value
#                 print(value)
#                 datetimeobject = datetime.strptime(value, '%Y%m%d')
#                 newformat2 = datetimeobject.strptime('%m/%d/%Y')
#                 print(newformat2)

# text = "3PI #25275 Project Based - 98313 - DPS - Workplace Experience (RWE) Program-DPS Raynham"
# text1 = text.split(" ").pop(1).split("#").pop(1)
# print(text1)

# import openpyxl
# path = "C:\\Users\\satheeshnair\\PycharmProjects\\3PI\\testdata\\testdata.xlsx"
# wb = openpyxl.load_workbook(path)
# sheet = wb.active
# c1 = sheet.cell(row=21, column=3).value = "satheesh"
# # c1.value = "satheesh"
# wb.save(path)

# path = "C:\\Users\\satheeshnair\\PycharmProjects\\3PI\\testdata\\testdata.xlsx"
# wb = openpyxl.load_workbook(path)
# sheet = wb.active
# rows = sheet.max_row
# cols = sheet.
# max_column
# for col in range(1, cols + 1):
#     for row in range(2, rows + 1):
#         testcasename = sheet.cell(row=row, column=col).value
#         if (testcasename == "test_review"):
#             methodname = sheet.cell(row=row, column=col + 1).value
#             if (methodname == "Order_Number"):
#                 review = sheet.cell(row=row, column=col + 2)
#                 review.value = str(9812)
#                 wb.save(path)

# print(os.getcwd())
# print(os.path.dirname (os.path.abspath("C:\\Users\\satheeshnair\\Desktop\\infocampus\\SourceCode\\AppianJJ\\configuration\\Drivers\\chromedriver.exe")))
# print(os.path.basename("C:\\Users\\satheeshnair\\Desktop\\infocampus\\SourceCode\\AppianJJ\\configuration\\Drivers\\
#
# print(os.path.abspath("driver\chromedriver.exe"))
# print(os.path.realpath("output/Pass"))


# import datetime, xlrd
# path=(r"C:\Users\satheeshnair\PycharmProjects\3PI\testdata\Testdata.xlsx")
# book = xlrd.open_workbook(path)
# sh = book.sheet_by_index(0)
# rows = sheet.max_row
# cols = sheet.max_column
# for col in range(1, cols + 1):
#     for row in range(2, rows + 1):
#         testcasename = sheet.cell(row=row, column=col).value
#         if (testcasename == "test_shipto"):
#             methodname = sheet.cell(row=row, column=col + 1).value
#             if (methodname == "Expected_date"):
#                 review = sheet.cell(row=row, column=col + 2).value
#                 print(review)
#                 # review.value = str(9812)
#                 # book.save(path)
#                 a1_as_datetime = datetime.datetime(*xlrd.xldate_as_tuple(review, book.datemode))
#                 print(a1_as_datetime)
#                 print ('datetime: %s' % a1_as_datetime)