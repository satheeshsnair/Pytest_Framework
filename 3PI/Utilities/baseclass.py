import inspect
import logging
import openpyxl
import pytest
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select


@pytest.mark.usefixtures("setup")
class Baseclass:

    def verifylinkpresent(self,text):
        element = WebDriverWait(self.driver,10).until(EC.presence_of_element_located((By.LINK_TEXT,text)))

    def selectOptionByText(self,locator,text):
        select = Select(locator)
        # select = Select(WebDriverWait(self.driver,10).until(EC.presence_of_element_located(locator)))
        select.select_by_visible_text(text)

    def waits(self):
        time.sleep(4)

    def getLogger(self):
        loggername = inspect.stack()[1][3]
        logger = logging.getLogger(loggername)
        filehandler = logging.FileHandler('logfile.log')
        formatter = logging.Formatter("%(asctime)s :%(levelname)s :%(name)s :%(message)s")
        filehandler.setFormatter(formatter)
        logger.addHandler(filehandler)
        logger.setLevel(logging.DEBUG)
        # logger.debug("A debug statement")
        # logger.info("Info")
        # logger.warning("Warning message")
        # logger.error("Test failed")
        # logger.critical("crictical error")
        return logger

    def getdata(self,Testcasename,Methodname):
        book = openpyxl.load_workbook("C:\\Users\\satheeshnair\\PycharmProjects\\3PI\\Testdata\\Testdata.xlsx")
        sheet = book.active
        rows = sheet.max_row
        cols = sheet.max_column

        for col in range(1, cols + 1):
            for row in range(2, rows + 1):
                testcasename = sheet.cell(row=row, column=col).value
                if (testcasename == Testcasename):
                    methodname = sheet.cell(row=row, column=col + 1).value
                    if (methodname == Methodname):
                        value = sheet.cell(row=row, column=col + 2).value
                        return(value)