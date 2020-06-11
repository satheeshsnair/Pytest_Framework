import inspect
import logging
import os
import openpyxl
import pytest
import time
import pyautogui

from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select


@pytest.mark.usefixtures("setup")
class Baseclass:

    def verifylinkpresent(self, text):
        element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.LINK_TEXT, text)))

    def selectOptionByText(self, locator, text):
        select = Select(locator)
        # select = Select(WebDriverWait(self.driver,10).until(EC.presence_of_element_located(locator)))
        select.select_by_visible_text(text)

    def waits(self,wait_time):
        time.sleep(wait_time)

    def expli_wait(self,locator):
        wait = WebDriverWait(self.driver, 20)
        element = wait.until(EC.presence_of_element_located((By.XPATH, locator)))

    def getLogger(self):
        loggername = inspect.stack()[1][3]
        logger = logging.getLogger(loggername)
        filepath = os.path.realpath("logs")
        filehandler = logging.FileHandler(filepath + '\\logfile.log')
        formatter = logging.Formatter("%(asctime)s :%(levelname)s :%(name)s :%(message)s")
        filehandler.setFormatter(formatter)
        logger.addHandler(filehandler)
        logger.setLevel(logging.DEBUG)
        # logger.debug("A debug statement")
        # logger.info("Info")
        # logger.warning("Warning message")
        # logger.error("Test failed")
        # logger.critical("crictical error")
        return logger

    def getdata(self, Testcasename, Methodname):
        # book = openpyxl.load_workbook("C:\\Users\\satheeshnair\\PycharmProjects\\3PI\\testdata\\testdata.xlsx")
        book = openpyxl.load_workbook(os.path.realpath("testdata/Testdata.xlsx"))
        sheet = book.active
        rows = sheet.max_row
        cols = sheet.max_column

        for col in range(1, cols + 1):
            for row in range(2, rows + 1):
                testcasename = sheet.cell(row=row, column=col).value
                if (testcasename == Testcasename):
                    methodname = sheet.cell(row=row, column=col + 1).value
                    if (methodname == Methodname):
                        value = sheet.cell(row=row, column=col + 2).value
                        return (value)

    def clickbutton(self, locator):
        self.driver.find_element_by_xpath(locator).click()

    def get_list_data(self,locator,testcase,method):
        self.driver.find_element_by_xpath(locator).click()
        self.waits(2)
        self.driver.find_element_by_xpath(locator).send_keys(self.getdata(testcase, method))
        self.waits(2)
        actions = ActionChains(self.driver)
        actions.send_keys(Keys.ENTER)
        self.waits(0.5)
        actions.perform()
        self.waits(2)

    def uploadfile(self,locator,path):
        try:
            file = self.driver.find_element_by_xpath(locator)
            self.waits(2)
            file.send_keys(path)
            self.waits(2)
        except NoSuchElementException:
            print("Element not found")

    def write_to_excel(self,Testcasename,Methodname,order_number):
        # path = "C:\\Users\\satheeshnair\\PycharmProjects\\3PI\\testdata\\testdata.xlsx"
        path = os.path.realpath("testdata/Testdata.xlsx")
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        rows = sheet.max_row
        cols = sheet.max_column
        for col in range(1, cols + 1):
            for row in range(2, rows + 1):
                testcasename = sheet.cell(row=row, column=col).value
                if (testcasename == Testcasename):
                    methodname = sheet.cell(row=row, column=col + 1).value
                    if (methodname == Methodname):
                        review = sheet.cell(row=row, column=col + 2)
                        review.value = order_number
                        wb.save(path)
                        break

    def Pass_snaps(self,testname):
        filepath = os.path.realpath("output/Pass")
        myScreenshot = pyautogui.screenshot()
        myScreenshot.save(filepath + "\\" + testname + ".png")

    def Fail_snaps(self,testname):
        filepath = os.path.realpath("output/Fail")
        myScreenshot = pyautogui.screenshot()
        myScreenshot.save(filepath + "\\" + testname + ".png")