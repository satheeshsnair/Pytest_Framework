import openpyxl

workbook = openpyxl.load_workbook(r"C:\Users\satheeshnair\PycharmProjects\Appian_JNJ\Testdata.xlsx")
sheet = workbook.get_sheet_by_name("login")
rows = sheet.max_row
cols = sheet.max_column


class utils():
    def __init__(self,driver):
        self.driver = driver

    def getun(Testcasename,Methodname):
        for row in range(2,rows+1):
            for col in range(1,cols+1):
                testcasename = sheet.cell(row=row,column=col).value
                if (testcasename == Testcasename):
                    methodname = sheet.cell(row=row,column=col+1).value
                    if(methodname == Methodname):
                        value = sheet.cell(row=row,column=col+2).value
                        return(value)


    # if __name__ == "__main__":
    #     print(getun("satheesh","test"))
